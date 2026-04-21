#!/usr/bin/env python3
"""CSV Editor — browser-based, zero external dependencies.

Usage:
  python3 csv_editor.py              # start with blank sheet
  python3 csv_editor.py data.csv     # open a file directly
"""

import base64
import csv
import io
import json
import sys
import threading
import warnings
import ssl
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

try:
    import openpyxl
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False


# ── Shared state ──────────────────────────────────────────────────────────────

class State:
    def __init__(self):
        self.sheets       = [{'name': 'Sheet1', 'headers': [], 'rows': []}]
        self.active_sheet = 0
        self.filepath     = None
        self.filetype     = 'csv'

    @property
    def headers(self):
        return self.sheets[self.active_sheet]['headers']

    @headers.setter
    def headers(self, v):
        self.sheets[self.active_sheet]['headers'] = v

    @property
    def rows(self):
        return self.sheets[self.active_sheet]['rows']

    @rows.setter
    def rows(self, v):
        self.sheets[self.active_sheet]['rows'] = v

state = State()


# ── JSON helpers ──────────────────────────────────────────────────────────────

def _json_column_keys(headers):
    seen = {}
    keys = []
    for idx, header in enumerate(headers):
        base = (header or '').strip() or _excel_col_name(idx)
        count = seen.get(base, 0) + 1
        seen[base] = count
        keys.append(base if count == 1 else f'{base}_{count}')
    return keys


def _sheets_to_doc_json(sheets):
    out = []
    for sheet in sheets or []:
        headers = list(sheet.get('headers', []))
        rows = [list(r) for r in sheet.get('rows', [])]
        keys = _json_column_keys(headers)
        sheet_rows = []
        for row in rows:
            row_obj = {}
            for idx, key in enumerate(keys):
                row_obj[key] = row[idx] if idx < len(row) else ''
            sheet_rows.append(row_obj)
        out.append({
            'name': sheet.get('name', 'Sheet1'),
            'rows': sheet_rows,
        })
    return {'doc': {'sheets': out}}


def _doc_json_to_sheets(raw):
    data = json.loads(raw) if isinstance(raw, str) else raw
    if not isinstance(data, dict):
        raise ValueError('JSON must be an object with a doc root')
    doc = data.get('doc')
    if not isinstance(doc, dict):
        raise ValueError('JSON must contain a doc object')
    raw_sheets = doc.get('sheets')
    if raw_sheets is None:
        raw_sheets = doc.get('sheet')
    if isinstance(raw_sheets, dict):
        raw_sheets = [raw_sheets]
    if raw_sheets is None:
        raw_sheets = []
    if not isinstance(raw_sheets, list):
        raise ValueError('doc.sheets must be a list')

    sheets = []
    for idx, raw_sheet in enumerate(raw_sheets):
        if not isinstance(raw_sheet, dict):
            continue
        name = raw_sheet.get('name') or f'Sheet{idx + 1}'
        raw_rows = raw_sheet.get('rows', [])
        if not isinstance(raw_rows, list):
            raise ValueError(f'rows for sheet "{name}" must be a list')
        headers = []
        for row in raw_rows:
            if not isinstance(row, dict):
                raise ValueError(f'rows for sheet "{name}" must contain objects')
            for key in row.keys():
                if key not in headers:
                    headers.append(str(key))
        rows = []
        for row in raw_rows:
            rows.append([_cell_to_str(row.get(header, '')) for header in headers])
        sheets.append({'name': name, 'headers': headers, 'rows': rows})
    return sheets or [{'name': 'Sheet1', 'headers': [], 'rows': []}]


# ── Excel helpers ─────────────────────────────────────────────────────────────

import datetime as _dt
import re as _re
import colorsys as _colorsys
import xml.etree.ElementTree as _ET
import zipfile as _zipfile

# Register XLSX namespaces so ElementTree serialises them with their original prefixes
# (without this, Python re-serialises the default namespace as ns0: which Excel rejects)
for _ns_prefix, _ns_uri in [
    ('',     'http://schemas.openxmlformats.org/spreadsheetml/2006/main'),
    ('r',    'http://schemas.openxmlformats.org/officeDocument/2006/relationships'),
    ('mc',   'http://schemas.openxmlformats.org/markup-compatibility/2006'),
    ('x14ac','http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac'),
    ('xr',   'http://schemas.microsoft.com/office/spreadsheetml/2014/revision'),
    ('xr2',  'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2'),
    ('xr3',  'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3'),
]:
    _ET.register_namespace(_ns_prefix, _ns_uri)
import posixpath as _ppath

def _cell_to_str(value):
    """Convert an openpyxl cell value to a clean string."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float):
        # Avoid ugly "984.0" for whole numbers
        return str(int(value)) if value == int(value) else str(value)
    if isinstance(value, _dt.datetime):
        # Strip the time portion when it is midnight (date-only cells)
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime('%Y-%m-%d')
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, _dt.date):
        return value.strftime('%Y-%m-%d')
    return str(value)


def _cell_to_str_formatted(value, number_format):
    """Convert a cell value to string, respecting the Excel number format."""
    # Dates/datetimes are already formatted by type in _cell_to_str
    if value is None or isinstance(value, bool) or isinstance(value, (_dt.date, _dt.datetime)):
        return _cell_to_str(value)
    if not isinstance(value, (int, float)):
        return _cell_to_str(value)
    fmt = number_format or 'General'
    if fmt in ('General', '@', ''):
        return _cell_to_str(value)
    # Use the first segment of compound formats (e.g. positive;negative;zero)
    fmt_part = fmt.split(';')[0]
    is_pct = '%' in fmt_part
    num = float(value) * 100 if is_pct else float(value)
    # Count decimal places: digits after the first '.' in the format
    m = _re.search(r'\.([0#]+)', fmt_part)
    dp = len(m.group(1)) if m else 0
    result = f'{num:.{dp}f}' if dp > 0 else str(int(round(num)))
    return result + '%' if is_pct else result

_INT_RE  = _re.compile(r'^-?\d+$')
_FLOAT_RE = _re.compile(r'^-?\d+\.?\d*([eE][+-]?\d+)?$')

def _str_to_cell(s):
    """Convert a string cell value back to an appropriate Python type for Excel."""
    if not isinstance(s, str):
        return s
    if _INT_RE.match(s):
        return int(s)
    if _FLOAT_RE.match(s):
        return float(s)
    return s


_THEME_ORDER = ['lt1','dk1','lt2','dk2','accent1','accent2','accent3','accent4','accent5','accent6','hlink','folHlink']
_INDEXED_COLORS = {
    0: '000000', 1: 'FFFFFF', 2: 'FF0000', 3: '00FF00', 4: '0000FF', 5: 'FFFF00', 6: 'FF00FF', 7: '00FFFF',
    8: '000000', 9: 'FFFFFF', 10: 'FF0000', 11: '00FF00', 12: '0000FF', 13: 'FFFF00', 14: 'FF00FF', 15: '00FFFF',
    16: '800000', 17: '008000', 18: '000080', 19: '808000', 20: '800080', 21: '008080', 22: 'C0C0C0', 23: '808080',
    24: '9999FF', 25: '993366', 26: 'FFFFCC', 27: 'CCFFFF', 28: '660066', 29: 'FF8080', 30: '0066CC', 31: 'CCCCFF',
    32: '000080', 33: 'FF00FF', 34: 'FFFF00', 35: '00FFFF', 36: '800080', 37: '800000', 38: '008080', 39: '0000FF',
    40: '00CCFF', 41: 'CCFFFF', 42: 'CCFFCC', 43: 'FFFF99', 44: '99CCFF', 45: 'FF99CC', 46: 'CC99FF', 47: 'FFCC99',
    48: '3366FF', 49: '33CCCC', 50: '99CC00', 51: 'FFCC00', 52: 'FF9900', 53: 'FF6600', 54: '666699', 55: '969696',
    56: '003366', 57: '339966', 58: '003300', 59: '333300', 60: '993300', 61: '993366', 62: '333399', 63: '333333',
}


def _parse_theme_colors(wb):
    xml = getattr(wb, 'loaded_theme', None)
    if not xml:
        return None
    try:
        root = _ET.fromstring(xml)
        scheme = root.find('.//{*}clrScheme')
        if scheme is None:
            return None
        colors = []
        for tag in _THEME_ORDER:
            el = scheme.find(f'{{*}}{tag}')
            if el is None:
                colors.append('000000')
                continue
            srgb = el.find('{*}srgbClr')
            sysc = el.find('{*}sysClr')
            colors.append(
                (srgb.get('val') if srgb is not None else None)
                or (sysc.get('lastClr') if sysc is not None else None)
                or '000000'
            )
        return colors
    except Exception:
        return None


def _apply_tint(hex_color, tint):
    if not tint:
        return hex_color.upper()
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0
    h, l, s = _colorsys.rgb_to_hls(r, g, b)
    l = l * (1 - tint) + tint if tint >= 0 else l * (1 + tint)
    l = max(0.0, min(1.0, l))
    nr, ng, nb = _colorsys.hls_to_rgb(h, l, s)
    return ''.join(f'{round(v * 255):02X}' for v in (nr, ng, nb))


def _resolve_openpyxl_color(color, theme_colors):
    if color is None:
        return None
    typ = getattr(color, 'type', None)
    if typ == 'rgb' and getattr(color, 'rgb', None):
        return '#' + color.rgb[-6:]
    if typ == 'theme' and getattr(color, 'theme', None) is not None and theme_colors:
        base = theme_colors[color.theme] if color.theme < len(theme_colors) else '000000'
        return '#' + _apply_tint(base, getattr(color, 'tint', 0) or 0)
    if typ == 'indexed' and getattr(color, 'indexed', None) in _INDEXED_COLORS:
        return '#' + _INDEXED_COLORS[color.indexed]
    return None


def _excel_col_name(idx):
    name = ''
    n = idx + 1
    while n:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def _row_to_visible_area(row_idx, header_idx):
    if header_idx < 0:
        return ('body', row_idx)
    if row_idx == header_idx:
        return ('header', 0)
    body_idx = row_idx
    if row_idx > header_idx:
        body_idx -= 1
    return ('body', body_idx)


def _parse_sheet_merges(ws, header_idx):
    merges = []
    for merged in ws.merged_cells.ranges:
        min_row = merged.min_row - 1
        max_row = merged.max_row - 1
        start_area, start_visible_row = _row_to_visible_area(min_row, header_idx)
        end_area, end_visible_row = _row_to_visible_area(max_row, header_idx)
        if not start_area or start_area != end_area:
            continue
        if start_area == 'header':
            if merged.min_row != merged.max_row:
                continue
            merges.append({
                'area': 'header',
                'row': 0,
                'col': merged.min_col - 1,
                'rowspan': 1,
                'colspan': merged.max_col - merged.min_col + 1,
            })
            continue
        merges.append({
            'area': 'body',
            'row': start_visible_row,
            'col': merged.min_col - 1,
            'rowspan': end_visible_row - start_visible_row + 1,
            'colspan': merged.max_col - merged.min_col + 1,
        })
    return merges


def _merge_ref_for_sheet(sheet, merge):
    area = merge.get('area', 'body')
    row = int(merge.get('row', 0) or 0)
    col = int(merge.get('col', 0) or 0)
    rowspan = max(1, int(merge.get('rowspan', 1) or 1))
    colspan = max(1, int(merge.get('colspan', 1) or 1))
    header_idx_raw = sheet.get('headerRowIndex', 0)
    header_idx = int(header_idx_raw) if header_idx_raw not in (None, '') else 0
    promoted_header = header_idx >= 0
    if area == 'header' and promoted_header:
        start_row = header_idx + 1
    else:
        if not promoted_header:
            start_row = row + 1
        else:
            start_row = row + 1 if row < header_idx else row + 2
    end_row = start_row + rowspan - 1
    start_col = col + 1
    end_col = start_col + colspan - 1
    return f'{_excel_col_name(start_col - 1)}{start_row}:{_excel_col_name(end_col - 1)}{end_row}'


def _propagate_consistent_column_alignment(rows2d, styles2d, header_idx):
    if not rows2d or not styles2d:
        return
    data_start = (header_idx + 1) if header_idx is not None and header_idx >= 0 else 0
    max_col = max((len(row) for row in rows2d), default=0)
    for col in range(max_col):
        aligned_rows = []
        align_values = set()
        last_nonempty = None
        for row in range(data_start, len(rows2d)):
            value = rows2d[row][col] if col < len(rows2d[row]) else ''
            if value == '':
                continue
            last_nonempty = row
            st = styles2d[row][col] if col < len(styles2d[row]) else None
            align = st.get('align') if st else None
            if align:
                aligned_rows.append(row)
                align_values.add(align)
        if len(aligned_rows) < 5 or len(align_values) != 1 or last_nonempty is None:
            continue
        last_aligned = aligned_rows[-1]
        if last_aligned >= last_nonempty:
            continue
        inferred_align = next(iter(align_values))
        for row in range(last_aligned + 1, last_nonempty + 1):
            value = rows2d[row][col] if col < len(rows2d[row]) else ''
            if value == '':
                continue
            while len(styles2d[row]) <= col:
                styles2d[row].append(None)
            st = dict(styles2d[row][col] or {})
            if st.get('align'):
                continue
            st['align'] = inferred_align
            styles2d[row][col] = st


def _sheet_row_to_body_row(sheet_row_idx, header_idx):
    if header_idx is None or header_idx < 0:
        return sheet_row_idx
    if sheet_row_idx < header_idx:
        return sheet_row_idx
    if sheet_row_idx > header_idx:
        return sheet_row_idx - 1
    return None


def _dxf_to_style(dxf, theme_colors):
    if not dxf:
        return None
    st = {}
    fill = getattr(dxf, 'fill', None)
    if fill and getattr(fill, 'patternType', None) != 'none':
        fg = _resolve_openpyxl_color(getattr(fill, 'fgColor', None), theme_colors)
        bg_fallback = _resolve_openpyxl_color(getattr(fill, 'bgColor', None), theme_colors)
        bg = bg_fallback if getattr(fill, 'patternType', None) in (None, '') else (fg or bg_fallback)
        if bg == '#000000' and bg_fallback:
            bg = bg_fallback
        if bg:
            st['bg'] = bg
    font = getattr(dxf, 'font', None)
    if font:
        fc = _resolve_openpyxl_color(getattr(font, 'color', None), theme_colors)
        if fc:
            st['color'] = fc
        if getattr(font, 'bold', False):
            st['bold'] = True
        if getattr(font, 'italic', False):
            st['italic'] = True
        if getattr(font, 'strike', False):
            st['strike'] = True
    return st or None


def _parse_conditional_formats(ws, header_idx, theme_colors):
    out = []
    cf_rules = getattr(getattr(ws, 'conditional_formatting', None), '_cf_rules', {}) or {}
    for sqref, rules in cf_rules.items():
        try:
            ranges = str(getattr(sqref, 'sqref', sqref)).split()
        except Exception:
            ranges = [str(sqref)]
        for range_ref in ranges:
            try:
                min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_ref)
            except Exception:
                continue
            body_row_start = _sheet_row_to_body_row(min_row - 1, header_idx)
            body_row_end = _sheet_row_to_body_row(max_row - 1, header_idx)
            if body_row_start is None or body_row_end is None:
                continue
            for rule in rules:
                if getattr(rule, 'type', None) != 'top10':
                    continue
                style = _dxf_to_style(getattr(rule, 'dxf', None), theme_colors)
                if not style:
                    continue
                out.append({
                    'type': 'top10',
                    'bottom': bool(getattr(rule, 'bottom', False)),
                    'rank': int(getattr(rule, 'rank', 10) or 10),
                    'rowStart': body_row_start,
                    'rowEnd': body_row_end,
                    'colStart': min_col - 1,
                    'colEnd': max_col - 1,
                    'style': style,
                })
    return out


def _parse_xlsx_sheets_with_styles(raw_bytes):
    if not _XLSX_OK:
        raise RuntimeError('openpyxl not installed — run: pip install openpyxl')
    with warnings.catch_warnings():
        warnings.filterwarnings(
            'ignore',
            message='Slicer List extension is not supported and will be removed',
            category=UserWarning,
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    theme_colors = _parse_theme_colors(wb)
    image_map = _parse_xlsx_images(raw_bytes)
    sheets = []
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        default_col_width = getattr(getattr(ws, 'sheet_format', None), 'defaultColWidth', None)
        col_widths = []
        for c in range(1, max_col + 1):
            dim = ws.column_dimensions.get(_excel_col_name(c - 1))
            width = getattr(dim, 'width', None) if dim else None
            if width is None:
                width = default_col_width
            col_widths.append(float(width) if width else None)
        rows2d = []
        styles2d = []
        for r in range(1, max_row + 1):
            row_vals = []
            row_styles = []
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                if cell.data_type == 'f':
                    v = cell.value
                    f = v.text if hasattr(v, 'text') else (v or '')
                    row_vals.append(_re.sub(r'_xlfn\.', '', f))
                else:
                    row_vals.append(_cell_to_str_formatted(cell.value, cell.number_format))
                st = {}
                if cell.number_format and cell.number_format not in ('General', '@', ''):
                    st['numFmt'] = cell.number_format
                fill = cell.fill
                font = cell.font
                align = cell.alignment
                if getattr(fill, 'fill_type', None) and fill.fill_type != 'none':
                    bg = _resolve_openpyxl_color(fill.fgColor, theme_colors) or _resolve_openpyxl_color(fill.bgColor, theme_colors)
                    if bg:
                        st['bg'] = bg
                fc = _resolve_openpyxl_color(getattr(font, 'color', None), theme_colors)
                if fc:
                    st['color'] = fc
                if getattr(font, 'bold', False):
                    st['bold'] = True
                if getattr(font, 'italic', False):
                    st['italic'] = True
                if getattr(font, 'strike', False):
                    st['strike'] = True
                if getattr(font, 'sz', None):
                    st['fontSize'] = font.sz
                if getattr(font, 'name', None):
                    st['fontFamily'] = font.name
                if getattr(align, 'horizontal', None):
                    st['align'] = align.horizontal
                border = cell.border
                for side, key in [('top', 'borderTop'), ('bottom', 'borderBottom'),
                                   ('left', 'borderLeft'), ('right', 'borderRight')]:
                    b = getattr(border, side, None)
                    bs = getattr(b, 'border_style', None) if b else None
                    if bs:
                        bc = _resolve_openpyxl_color(getattr(b, 'color', None), theme_colors) or '#000000'
                        st[key] = {'style': bs, 'color': bc}
                row_styles.append(st or None)
            rows2d.append(row_vals)
            styles2d.append(row_styles)
        while rows2d and all(v == '' for v in rows2d[-1]) and all(s is None for s in styles2d[-1]):
            rows2d.pop()
            styles2d.pop()
        header_idx = None
        if ws.auto_filter and ws.auto_filter.ref:
            try:
                header_idx = openpyxl.utils.range_boundaries(ws.auto_filter.ref)[1] - 1
            except Exception:
                header_idx = None
        if header_idx is None and ws.tables:
            try:
                first_table = next(iter(ws.tables.values()))
                header_idx = openpyxl.utils.range_boundaries(first_table.ref)[1] - 1
            except Exception:
                header_idx = None

        _propagate_consistent_column_alignment(rows2d, styles2d, header_idx)

        if header_idx is None:
            headers = [_excel_col_name(i) for i in range(max_col)]
            header_styles = []
            leading_rows = rows2d
            trailing_rows = []
            leading_styles = styles2d
            trailing_styles = []
            header_idx_out = -1
        else:
            if not rows2d:
                header_idx = 0
            header_idx = max(0, min(header_idx, len(rows2d) - 1)) if rows2d else 0
            headers = rows2d[header_idx] if rows2d else []
            header_styles = styles2d[header_idx] if styles2d else []
            leading_rows = rows2d[:header_idx]
            trailing_rows = rows2d[header_idx + 1:] if len(rows2d) > header_idx + 1 else []
            leading_styles = styles2d[:header_idx]
            trailing_styles = styles2d[header_idx + 1:] if len(styles2d) > header_idx + 1 else []
            header_idx_out = header_idx
        sheets.append({
            'name': ws.title,
            'headers': headers,
            'rows': leading_rows + trailing_rows,
            'headerStyles': header_styles,
            'rowStyles': leading_styles + trailing_styles,
            'images': image_map.get(ws.title, []),
            'headerRowIndex': header_idx_out,
            'columnWidths': col_widths,
            'conditionalFormats': _parse_conditional_formats(ws, header_idx_out, theme_colors),
            'merges': _parse_sheet_merges(ws, header_idx_out),
        })
    return sheets


def _ordered_sheet_rows(sheet):
    header_idx = int(sheet.get('headerRowIndex', 0) or 0)
    if header_idx < 0:
        return [list(r) for r in sheet.get('rows', [])]
    rows = list(sheet.get('rows', []))
    header = list(sheet.get('headers', []))
    leading = rows[:header_idx]
    trailing = rows[header_idx:]
    return leading + [header] + trailing


def _load_shared_strings_from_zip(src_zip):
    """Returns (ss_list, ss_dict) preserving all original entries including duplicates."""
    if 'xl/sharedStrings.xml' not in src_zip.namelist():
        return [], {}
    try:
        root = _ET.fromstring(src_zip.read('xl/sharedStrings.xml'))
        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        ss_list = []
        ss_dict = {}
        for si in root.findall(f'{{{ns}}}si'):
            text = ''.join(t.text or '' for t in si.iter(f'{{{ns}}}t'))
            ss_list.append(text)
            if text not in ss_dict:
                ss_dict[text] = len(ss_list) - 1
        return ss_list, ss_dict
    except Exception:
        return [], {}


def _append_new_strings_to_ss_xml(orig_data, new_texts):
    """Append new <si> entries to existing sharedStrings.xml bytes."""
    import re as _re
    try:
        s = orig_data.decode('utf-8')
        parts = []
        for text in new_texts:
            escaped = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            sp = ' xml:space="preserve"' if (text.startswith(' ') or text.endswith(' ') or '\n' in text) else ''
            parts.append(f'<si><t{sp}>{escaped}</t></si>')
        n = len(new_texts)
        s = s.replace('</sst>', ''.join(parts) + '</sst>')
        s = _re.sub(r'\bcount="(\d+)"',       lambda m: f'count="{int(m.group(1)) + n}"',       s)
        s = _re.sub(r'\buniqueCount="(\d+)"', lambda m: f'uniqueCount="{int(m.group(1)) + n}"', s)
        return s.encode('utf-8')
    except Exception:
        return orig_data


def _set_xml_cell_value(cell_el, value, shared_strings=None, ss_list=None):
    original_t = cell_el.attrib.get('t')
    original_formula = next((child for child in list(cell_el) if child.tag.endswith('f')), None)
    original_formula_attrs = dict(original_formula.attrib) if original_formula is not None else None
    for child in list(cell_el):
        if child.tag.endswith(('v', 'is', 'f')):
            cell_el.remove(child)
    if value == '' or value is None:
        cell_el.attrib.pop('t', None)
        return
    if isinstance(value, str) and value.startswith('='):
        cell_el.attrib.pop('t', None)
        f = _ET.SubElement(cell_el, cell_el.tag[:-1] + 'f')
        if original_formula_attrs:
            # Strip shared/array formula metadata — having explicit formula text
            # alongside t="shared"/si is invalid OOXML and triggers repair warnings
            safe = {k: v for k, v in original_formula_attrs.items()
                    if k not in ('si',) and not (k == 't' and v in ('shared',))}
            if safe:
                f.attrib.update(safe)
        f.text = value[1:]
        return
    py_val = _str_to_cell(value) if isinstance(value, str) else value
    force_string = original_t in ('s', 'str', 'inlineStr')
    if isinstance(value, str):
        stripped = value.strip()
        has_ambiguous_leading_zero = (
            stripped.startswith('0')
            and len(stripped) > 1
            and stripped[1:].isdigit()
        ) or (
            stripped.startswith('-0')
            and len(stripped) > 2
            and stripped[2:].isdigit()
        )
        if has_ambiguous_leading_zero:
            force_string = True
    if isinstance(py_val, bool):
        cell_el.set('t', 'b')
        v = _ET.SubElement(cell_el, cell_el.tag[:-1] + 'v')
        v.text = '1' if py_val else '0'
        return
    if isinstance(py_val, (int, float)) and not force_string:
        cell_el.attrib.pop('t', None)
        v = _ET.SubElement(cell_el, cell_el.tag[:-1] + 'v')
        v.text = str(py_val)
        return
    str_val = '' if value is None else str(value)
    if shared_strings is not None:
        idx = shared_strings.get(str_val)
        if idx is not None:
            cell_el.set('t', 's')
            v = _ET.SubElement(cell_el, cell_el.tag[:-1] + 'v')
            v.text = str(idx)
            return
    cell_el.set('t', 'inlineStr')
    is_el = _ET.SubElement(cell_el, cell_el.tag[:-1] + 'is')
    t_el = _ET.SubElement(is_el, cell_el.tag[:-1] + 't')
    if isinstance(value, str) and (value.startswith(' ') or value.endswith(' ') or '\n' in value):
        t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    t_el.text = str_val


def _apply_clear_colors_to_styles_xml(data, cleared_bg_hex, cleared_text_hex):
    import re as _re
    if not cleared_bg_hex and not cleared_text_hex:
        return data
    try:
        s = data.decode('utf-8')
        if cleared_bg_hex:
            def _clear_fill(m):
                rgb_m = _re.search(r'rgb="[A-Fa-f0-9]{0,2}([A-Fa-f0-9]{6})"', m.group(0))
                if rgb_m and rgb_m.group(1).lower() in cleared_bg_hex:
                    return '<fill><patternFill patternType="none"/></fill>'
                return m.group(0)
            s = _re.sub(
                r'<fill>\s*<patternFill\s+patternType="solid">.*?</patternFill>\s*</fill>',
                _clear_fill, s, flags=_re.DOTALL)
        return s.encode('utf-8')
    except Exception:
        return data


def _write_xlsx_from_template(raw_bytes, sheets, cleared_bg=None, cleared_text=None):
    import re as _re
    _cleared_bg_hex  = {c.lstrip('#').lower()[-6:] for c in (cleared_bg  or []) if c}
    _cleared_txt_hex = {c.lstrip('#').lower()[-6:] for c in (cleared_text or []) if c}
    ns = {
        'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }
    main_ns_uri = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    # Register as default namespace so tostring() emits plain <sheetData> not <ns0:sheetData>
    _ET.register_namespace('', main_ns_uri)
    _ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
    by_name = {sh.get('name', ''): sh for sh in sheets}
    zin = io.BytesIO(raw_bytes)
    zout = io.BytesIO()
    with _zipfile.ZipFile(zin, 'r') as src, _zipfile.ZipFile(zout, 'w', _zipfile.ZIP_DEFLATED) as dst:
        wb = _ET.fromstring(src.read('xl/workbook.xml'))
        wb_rels = _ET.fromstring(src.read('xl/_rels/workbook.xml.rels'))
        rel_targets = {rel.attrib['Id']: rel.attrib['Target'] for rel in wb_rels.findall('{*}Relationship')}
        has_ss_file = 'xl/sharedStrings.xml' in src.namelist()
        _ss_list, shared_strings = _load_shared_strings_from_zip(src)
        # Only use t="s" references when the original file has a sharedStrings table.
        ss_ref = shared_strings if has_ss_file else None
        sheet_paths = {}
        for sheet in wb.findall('a:sheets/a:sheet', ns):
            rid = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            target = rel_targets.get(rid)
            if target:
                sheet_paths[sheet.attrib.get('name', '')] = _ppath.normpath('xl/' + target)

        for name in src.namelist():
            data = src.read(name)
            if name == 'xl/styles.xml' and (_cleared_bg_hex or _cleared_txt_hex):
                data = _apply_clear_colors_to_styles_xml(data, _cleared_bg_hex, _cleared_txt_hex)
            if name in sheet_paths.values():
                sheet_name = next((k for k, v in sheet_paths.items() if v == name), None)
                sheet = by_name.get(sheet_name or '')
                if sheet:
                    root = _ET.fromstring(data)
                    sheet_data_el = root.find('a:sheetData', ns)
                    if sheet_data_el is not None:
                        rows = _ordered_sheet_rows(sheet)
                        max_row = max(len(rows), len(sheet_data_el.findall('a:row', ns)))
                        max_col = max([len(r) for r in rows] + [0])
                        row_map = {
                            int(row.attrib.get('r', idx + 1)): row
                            for idx, row in enumerate(sheet_data_el.findall('a:row', ns))
                        }
                        cell_maps = {}
                        for rnum, row_el in row_map.items():
                            cmap = {}
                            for cell in row_el.findall('a:c', ns):
                                ref = cell.attrib.get('r', '')
                                col = ''.join(ch for ch in ref if ch.isalpha())
                                cmap[col] = cell
                            cell_maps[rnum] = cmap

                        # Build new sheetData element in memory
                        new_sd = _ET.Element(f'{{{main_ns_uri}}}sheetData')
                        for r in range(1, max_row + 1):
                            old_row = row_map.get(r)
                            row_el = _ET.Element(old_row.tag if old_row is not None else f'{{{main_ns_uri}}}row')
                            if old_row is not None:
                                row_el.attrib.update(old_row.attrib)
                            row_el.set('r', str(r))
                            values = rows[r - 1] if r - 1 < len(rows) else []
                            old_cells = cell_maps.get(r, {})
                            for c in range(max(max_col, len(values))):
                                col_name = _excel_col_name(c)
                                ref = f'{col_name}{r}'
                                old_cell = old_cells.get(col_name)
                                val = values[c] if c < len(values) else ''
                                if old_cell is None and (val == '' or val is None):
                                    continue
                                cell_el = _ET.fromstring(_ET.tostring(old_cell)) if old_cell is not None else _ET.Element(f'{{{main_ns_uri}}}c')
                                cell_el.set('r', ref)
                                _set_xml_cell_value(cell_el, val, shared_strings=ss_ref)
                                row_el.append(cell_el)
                            new_sd.append(row_el)

                        # Serialize sheetData only; strip namespace declarations since root already has them
                        new_sd_str = _ET.tostring(new_sd, encoding='unicode')
                        new_sd_str = _re.sub(r' xmlns(?::\w+)?="[^"]*"', '', new_sd_str)

                        # Operate on original XML bytes to preserve root element, namespaces, and attributes exactly
                        orig_str = data.decode('utf-8')

                        # Replace sheetData (handles both empty <sheetData/> and <sheetData>...</sheetData>)
                        _new_sd = new_sd_str  # closure for lambda
                        orig_str = _re.sub(
                            r'<(?:\w+:)?sheetData(?:\s[^>]*)?>.*?</(?:\w+:)?sheetData>|<(?:\w+:)?sheetData(?:\s[^>]*)?/>',
                            lambda m: _new_sd, orig_str, flags=_re.DOTALL)

                        # Update dimension ref
                        if max_row and max_col:
                            new_dim = f'A1:{_excel_col_name(max_col - 1)}{max_row}'
                            orig_str = _re.sub(
                                r'(<dimension\s+ref=")[^"]*(")',
                                lambda m: m.group(1) + new_dim + m.group(2),
                                orig_str)

                        # Build merge cells XML (no namespace prefix — inherits from root default ns)
                        merge_refs = []
                        for merge in sheet.get('merges', []) or []:
                            try:
                                merge_refs.append(_merge_ref_for_sheet(sheet, merge))
                            except Exception:
                                continue

                        # Remove existing mergeCells block
                        orig_str = _re.sub(r'<mergeCells(?:\s[^>]*)?>.*?</mergeCells>', '', orig_str, flags=_re.DOTALL)
                        orig_str = _re.sub(r'<mergeCells(?:\s[^>]*)?/>', '', orig_str)

                        if merge_refs:
                            mc_xml = f'<mergeCells count="{len(merge_refs)}">{"".join(f"<mergeCell ref=\"{ref}\"/>" for ref in merge_refs)}</mergeCells>'
                            orig_str = _re.sub(r'(</(?:\w+:)?sheetData>)', lambda m: m.group(1) + mc_xml, orig_str, count=1)

                        data = orig_str.encode('utf-8')
            dst.writestr(name, data)
    return zout.getvalue()


def _parse_xlsx_images(raw_bytes):
    ns = {
        'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'd': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'pr': 'http://schemas.openxmlformats.org/package/2006/relationships',
    }
    out = {}
    try:
        with _zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
            wb = _ET.fromstring(zf.read('xl/workbook.xml'))
            wb_rels = _ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            rel_targets = {rel.attrib['Id']: rel.attrib['Target'] for rel in wb_rels.findall('{*}Relationship')}
            for sheet in wb.findall('a:sheets/a:sheet', ns):
                name = sheet.attrib.get('name', '')
                rid = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                target = rel_targets.get(rid)
                if not target:
                    continue
                sheet_path = _ppath.normpath('xl/' + target)
                rels_path = _ppath.normpath(_ppath.join(_ppath.dirname(sheet_path), '_rels', _ppath.basename(sheet_path) + '.rels'))
                if rels_path not in zf.namelist():
                    continue
                sheet_rels = _ET.fromstring(zf.read(rels_path))
                sheet_rel_targets = {rel.attrib['Id']: rel.attrib['Target'] for rel in sheet_rels.findall('{*}Relationship')}
                sheet_xml = _ET.fromstring(zf.read(sheet_path))
                drawing = sheet_xml.find('a:drawing', ns)
                if drawing is None:
                    continue
                drawing_rid = drawing.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                drawing_target = sheet_rel_targets.get(drawing_rid)
                if not drawing_target:
                    continue
                drawing_path = _ppath.normpath(_ppath.join(_ppath.dirname(sheet_path), drawing_target))
                drawing_rels_path = _ppath.normpath(_ppath.join(_ppath.dirname(drawing_path), '_rels', _ppath.basename(drawing_path) + '.rels'))
                if drawing_path not in zf.namelist() or drawing_rels_path not in zf.namelist():
                    continue
                drawing_xml = _ET.fromstring(zf.read(drawing_path))
                drawing_rels = _ET.fromstring(zf.read(drawing_rels_path))
                drawing_rel_targets = {rel.attrib['Id']: rel.attrib['Target'] for rel in drawing_rels.findall('{*}Relationship')}
                images = []
                for anchor in drawing_xml.findall('xdr:twoCellAnchor', ns) + drawing_xml.findall('xdr:oneCellAnchor', ns):
                    blip = anchor.find('.//d:blip', ns)
                    if blip is None:
                        continue
                    embed = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    media_target = drawing_rel_targets.get(embed)
                    if not media_target:
                        continue
                    media_path = _ppath.normpath(_ppath.join(_ppath.dirname(drawing_path), media_target))
                    if media_path not in zf.namelist():
                        continue
                    data = zf.read(media_path)
                    ext = _ppath.splitext(media_path)[1].lower()
                    mime = {
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.png': 'image/png',
                        '.gif': 'image/gif',
                    }.get(ext, 'application/octet-stream')
                    fr = anchor.find('xdr:from', ns)
                    if fr is not None:
                        row = int(fr.findtext('xdr:row', default='0', namespaces=ns))
                        col = int(fr.findtext('xdr:col', default='0', namespaces=ns))
                    else:
                        row = col = 0
                    images.append({
                        'row': row,
                        'col': col,
                        'src': f'data:{mime};base64,' + base64.b64encode(data).decode('ascii'),
                    })
                if images:
                    out[name] = images
    except Exception:
        return {}
    return out


def _parse_xlsx(raw_bytes):
    """Parse an .xlsx binary into (headers, rows). Requires openpyxl."""
    if not _XLSX_OK:
        raise RuntimeError('openpyxl not installed — run: pip install openpyxl')
    with warnings.catch_warnings():
        warnings.filterwarnings(
            'ignore',
            message='Slicer List extension is not supported and will be removed',
            category=UserWarning,
        )
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    ws = wb.active
    if ws is None:
        # Workbook has no active sheet — try first sheet
        if not wb.sheetnames:
            return [], []
        ws = wb[wb.sheetnames[0]]
    all_rows = []
    for row in ws.iter_rows(values_only=False):
        cells = [_re.sub(r'_xlfn\.', '', v.text if hasattr(v := cell.value, 'text') else (v or '')) if cell.data_type == 'f' else _cell_to_str(cell.value) for cell in row]
        # Skip rows that are entirely blank
        if any(v != '' for v in cells):
            all_rows.append(cells)
    if not all_rows:
        return [], []
    n_cols = max(len(r) for r in all_rows)
    for r in all_rows:
        while len(r) < n_cols:
            r.append('')
    return all_rows[0], all_rows[1:]


def _write_xlsx(headers, rows):
    """Serialise headers + rows to .xlsx bytes. Requires openpyxl.
    Numeric-looking strings are written as numbers so Excel treats them correctly."""
    if not _XLSX_OK:
        raise RuntimeError('openpyxl not installed — run: pip install openpyxl')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)  # headers always as strings
    for row in rows:
        ws.append([_str_to_cell(c) for c in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── GitHub state & helpers ────────────────────────────────────────────────────

class GitHubConfig:
    def __init__(self):
        self.token  = ''
        self.repo   = ''    # owner/repo
        self.branch = 'main'
        self.path   = ''    # file path inside repo

gh = GitHubConfig()

_GH_CONFIG_FILE = Path.home() / '.csv_editor_github.json'

def load_gh_config():
    """Load persisted GitHub config from disk, if it exists."""
    try:
        data = json.loads(_GH_CONFIG_FILE.read_text(encoding='utf-8'))
        gh.token  = data.get('token',  '')
        gh.repo   = data.get('repo',   '')
        gh.branch = data.get('branch', 'main') or 'main'
        gh.path   = data.get('path',   '')
    except (FileNotFoundError, json.JSONDecodeError):
        pass

def save_gh_config():
    """Persist current GitHub config to disk."""
    _GH_CONFIG_FILE.write_text(
        json.dumps({'token': gh.token, 'repo': gh.repo, 'branch': gh.branch, 'path': gh.path}),
        encoding='utf-8',
    )


def _ssl_context():
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        pass
    return ssl._create_unverified_context()

def _gh_api(method, endpoint, data=None):
    """Call GitHub REST API. Returns (result_dict, error_str)."""
    url = 'https://api.github.com' + endpoint
    hdrs = {
        'Authorization': f'token {gh.token}',
        'Accept': 'application/vnd.github.v3+json',
        'User-Agent': 'CSVEditor/1.0',
    }
    body = json.dumps(data).encode() if data is not None else None
    if body:
        hdrs['Content-Type'] = 'application/json'
    req = urllib.request.Request(url, data=body, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, context=_ssl_context()) as r:
            return json.loads(r.read()), None
    except urllib.error.HTTPError as e:
        try:
            msg = json.loads(e.read()).get('message', str(e))
        except Exception:
            msg = str(e)
        return None, msg
    except Exception as e:
        return None, str(e)


# ── HTML/CSS/JS (loaded from index.html next to this script) ─────────────────

_HTML_FILE = Path(__file__).parent / 'index.html'

def _load_html():
    return _HTML_FILE.read_text(encoding='utf-8')


# ── HTTP handler ──────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):

    def do_GET(self):
        if self.path == '/':
            self._serve_html()
        elif self.path == '/api/data':
            active = state.sheets[state.active_sheet]
            self._json({
                'headers':     state.headers,
                'rows':        state.rows,
                'headerRowIndex': active.get('headerRowIndex', 0),
                'columnWidths': active.get('columnWidths', []),
                'filepath':    state.filepath,
                'filetype':    state.filetype,
                'sheets':      [{'name': s['name']} for s in state.sheets],
                'activeSheet': state.active_sheet,
            })
        elif self.path == '/api/all-sheets':
            self._json({'sheets': state.sheets, 'activeSheet': state.active_sheet})
        elif self.path == '/api/github/config':
            self._json({
                'token':       ('*' * 4 + gh.token[-4:]) if len(gh.token) > 4 else ('*' * len(gh.token)),
                'repo':        gh.repo,
                'branch':      gh.branch,
                'path':        gh.path,
                'configured':  bool(gh.token and gh.repo and gh.path),
            })
        elif self.path == '/api/github/history':
            if not gh.token or not gh.repo:
                self._json({'error': 'Not configured'}); return
            qs = f'?path={urllib.parse.quote(gh.path, safe="/")}&sha={urllib.parse.quote(gh.branch, safe="/")}&per_page=40'
            result, err = _gh_api('GET', f'/repos/{gh.repo}/commits{qs}')
            if err:
                self._json({'error': err}); return
            self._json({'commits': [{
                'sha':     c['sha'],
                'short':   c['sha'][:7],
                'message': c['commit']['message'].split('\n')[0][:80],
                'author':  c['commit']['author']['name'],
                'date':    c['commit']['author']['date'],
            } for c in result]})
        elif self.path.startswith('/api/github/version'):
            qs   = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            sha  = qs.get('sha', [''])[0]
            path = qs.get('path', [gh.path])[0]
            if not sha or not gh.token or not gh.repo:
                self._json({'error': 'Missing params'}); return
            result, err = _gh_api('GET', f'/repos/{gh.repo}/contents/{path.lstrip("/")}?ref={sha}')
            if err:
                self._json({'error': err}); return
            raw_bytes = base64.b64decode(result['content'].replace('\n', ''))
            # Detect xlsx by extension OR by zip magic bytes (PK\x03\x04) — gh.path may lack extension
            is_xlsx = path.lower().endswith('.xlsx') or raw_bytes[:4] == b'PK\x03\x04'
            if is_xlsx:
                try:
                    sheets = _parse_xlsx_sheets_with_styles(raw_bytes)
                    self._json({
                        'sheets': sheets,
                        'rawB64': base64.b64encode(raw_bytes).decode('ascii'),
                        'sha': sha,
                        'format': 'xlsx',
                    })
                except Exception:
                    # Fallback to raw bytes if server-side style parsing fails
                    self._json({'rawB64': base64.b64encode(raw_bytes).decode('ascii'), 'sha': sha, 'format': 'xlsx'})
            else:
                try:
                    content = raw_bytes.decode('utf-8-sig')
                except UnicodeDecodeError:
                    content = raw_bytes.decode('latin-1')
                self._json({'content': content, 'sha': sha, 'format': 'csv'})
        else:
            self._send(404, b'Not found')

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length) if length else b'{}'
        # Binary endpoints must be handled before JSON parsing
        if self.path.startswith('/api/github/commit-xlsx'):
            qs      = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            message = qs.get('message', ['Update file'])[0] or 'Update file'
            if not gh.token or not gh.repo or not gh.path:
                self._json({'error': 'GitHub not configured'}); return
            clean_path = gh.path.lstrip('/')
            existing, _ = _gh_api('GET',
                f'/repos/{gh.repo}/contents/{clean_path}?ref={urllib.parse.quote(gh.branch)}')
            file_sha = existing.get('sha') if isinstance(existing, dict) else None
            encoded  = base64.b64encode(body).decode('ascii')
            payload  = {'message': message, 'content': encoded, 'branch': gh.branch}
            if file_sha:
                payload['sha'] = file_sha
            result, err = _gh_api('PUT', f'/repos/{gh.repo}/contents/{clean_path}', payload)
            if err:
                self._json({'error': err})
            else:
                self._json({'ok': True, 'sha': result['commit']['sha'][:7],
                            'url': result['commit']['html_url']})
            return
        if self.path.startswith('/api/save-xlsx'):
            qs       = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            filepath = qs.get('filepath', [''])[0] or state.filepath or ''
            print(f'[save-xlsx] filepath={filepath!r} body_len={len(body)}')
            if not filepath:
                print('[save-xlsx] ERROR: no_path')
                self._json({'ok': False, 'error': 'no_path'}); return
            if '/' not in filepath and '\\' not in filepath:
                filepath = str(Path.cwd() / filepath)
                print(f'[save-xlsx] resolved to {filepath!r}')
            try:
                if not body:
                    print('[save-xlsx] ERROR: no data received')
                    self._json({'ok': False, 'error': 'no data received'}); return
                p = Path(filepath)
                p.parent.mkdir(parents=True, exist_ok=True)
                p.write_bytes(body)
                print(f'[save-xlsx] wrote {len(body)} bytes to {filepath!r}')
                state.filepath = filepath
                state.filetype = 'xlsx'
                self._json({'ok': True, 'filepath': filepath})
            except Exception as e:
                print(f'[save-xlsx] ERROR: {e}')
                self._json({'ok': False, 'error': str(e)})
            return
        if self.path == '/api/parse-xlsx':
            if not _XLSX_OK:
                self._json({'error': 'openpyxl not installed — run: pip install openpyxl'}); return
            try:
                sheets = _parse_xlsx_sheets_with_styles(body)
            except Exception as e:
                self._json({'error': str(e)}); return
            self._json({'ok': True, 'sheets': sheets})
            return
        if self.path.startswith('/api/build-xlsx-from-template'):
            qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            cleared_bg   = [c for c in qs.get('cleared_bg',    [''])[0].split(',') if c.strip()]
            cleared_text = [c for c in qs.get('cleared_color', [''])[0].split(',') if c.strip()]
            print(f'[build-xlsx-from-template v7] body={len(body)}B cleared_bg={cleared_bg}', flush=True)
            try:
                raw = _write_xlsx_from_template(body, state.sheets,
                                                cleared_bg=cleared_bg, cleared_text=cleared_text)
                print(f'[build-xlsx-from-template v7] OK output={len(raw)}B', flush=True)
            except Exception as e:
                import traceback as _tb
                print(f'[build-xlsx-from-template v7] ERROR:\n{_tb.format_exc()}', flush=True)
                self._json({'error': str(e)}); return
            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Length', str(len(raw)))
            self.end_headers()
            self.wfile.write(raw)
            return
        try:
            data = json.loads(body)
        except json.JSONDecodeError:
            data = {}

        if self.path == '/api/load-content':
            filename = data.get('filename', 'data.csv')
            fmt      = data.get('format', 'csv')
            if fmt == 'xlsx':
                # xlsx is parsed client-side by SheetJS; server receives {sheets: [{name, headers, rows}]}
                raw_sheets = data.get('sheets', [])
                if not raw_sheets:
                    raw_sheets = [{'name': 'Sheet1', 'headers': [], 'rows': []}]
                state.sheets       = raw_sheets
                state.active_sheet = 0
                state.filetype     = 'xlsx'
            elif fmt == 'json':
                content = data.get('content', '')
                try:
                    state.sheets = _doc_json_to_sheets(content)
                except Exception as e:
                    self._json({'ok': False, 'error': str(e)}); return
                state.active_sheet = 0
                state.filetype     = 'json'
            else:
                content  = data.get('content', '')
                reader   = csv.reader(io.StringIO(content))
                all_rows = list(reader)
                if all_rows:
                    state_headers = all_rows[0]
                    state_rows    = [list(r) for r in all_rows[1:]]
                    for row in state_rows:
                        while len(row) < len(state_headers):
                            row.append('')
                else:
                    state_headers = []
                    state_rows    = []
                state.sheets       = [{'name': 'Sheet1', 'headers': state_headers, 'rows': state_rows}]
                state.active_sheet = 0
                state.filetype     = 'csv'
            state.filepath = filename
            self._json({'ok': True})

        elif self.path == '/api/update':
            state.headers = data.get('headers', state.headers)
            state.rows    = data.get('rows',    state.rows)
            active = state.sheets[state.active_sheet]
            active['headerRowIndex'] = data.get('headerRowIndex', active.get('headerRowIndex', 0))
            active['columnWidths'] = data.get('columnWidths', active.get('columnWidths', []))
            active['headerStyles'] = data.get('headerStyles', active.get('headerStyles', []))
            active['rowStyles'] = data.get('rowStyles', active.get('rowStyles', []))
            active['merges'] = data.get('merges', active.get('merges', []))
            self._json({'ok': True})

        elif self.path == '/api/save':
            filepath = data.get('filepath') or state.filepath
            if not filepath:
                self._json({'ok': False, 'error': 'no_path'})
                return
            # If only a bare filename (no directory), resolve against cwd
            p_check = Path(filepath)
            if not p_check.is_absolute() and '/' not in filepath and '\\' not in filepath:
                filepath = str(Path.cwd() / filepath)
            try:
                p = Path(filepath)
                p.parent.mkdir(parents=True, exist_ok=True)
                if p.suffix.lower() == '.xlsx':
                    # xlsx bytes are generated client-side by SheetJS and written
                    # via /api/save-bytes — this branch handles CSV-only saves
                    # (fallback: if raw_bytes provided here, write them directly)
                    raw_b64 = data.get('raw_bytes')
                    if raw_b64:
                        p.write_bytes(base64.b64decode(raw_b64))
                    else:
                        self._json({'ok': False, 'error': 'xlsx_needs_bytes'}); return
                    state.filetype = 'xlsx'
                elif p.suffix.lower() == '.json':
                    p.write_text(json.dumps(_sheets_to_doc_json(state.sheets), ensure_ascii=False, indent=2), encoding='utf-8')
                    state.filetype = 'json'
                else:
                    with open(filepath, 'w', newline='', encoding='utf-8') as f:
                        csv.writer(f).writerow(state.headers)
                        csv.writer(f).writerows(state.rows)
                    state.filetype = 'csv'
                state.filepath = filepath
                self._json({'ok': True, 'filepath': filepath, 'filetype': state.filetype})
            except Exception as e:
                self._json({'ok': False, 'error': str(e)})

        elif self.path == '/api/export':
            fmt = data.get('format', state.filetype)
            if fmt == 'xlsx':
                if not _XLSX_OK:
                    self._json({'error': 'openpyxl not installed — run: pip install openpyxl'}); return
                try:
                    raw = _write_xlsx(state.headers, state.rows)
                except Exception as e:
                    self._json({'error': str(e)}); return
                self.send_response(200)
                self.send_header('Content-Type',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Length', str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
            elif fmt == 'json':
                raw = json.dumps(_sheets_to_doc_json(state.sheets), ensure_ascii=False, indent=2).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
            else:
                buf = io.StringIO()
                w   = csv.writer(buf)
                w.writerow(state.headers)
                w.writerows(state.rows)
                raw = buf.getvalue().encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'text/csv; charset=utf-8')
                self.send_header('Content-Length', str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)

        elif self.path == '/api/switch-sheet':
            idx = int(data.get('index', 0))
            if 0 <= idx < len(state.sheets):
                state.active_sheet = idx
            self._json({'ok': True})

        elif self.path == '/api/add-sheet':
            name = data.get('name', '').strip()
            if not name:
                existing = {s['name'] for s in state.sheets}
                i = len(state.sheets) + 1
                while f'Sheet{i}' in existing:
                    i += 1
                name = f'Sheet{i}'
            state.sheets.append({'name': name, 'headers': list(state.headers) if state.headers else [], 'rows': []})
            state.active_sheet = len(state.sheets) - 1
            self._json({'ok': True, 'index': state.active_sheet, 'name': name})

        elif self.path == '/api/rename-sheet':
            idx  = int(data.get('index', state.active_sheet))
            name = data.get('name', '').strip()
            if name and 0 <= idx < len(state.sheets):
                state.sheets[idx]['name'] = name
            self._json({'ok': True})

        elif self.path == '/api/delete-sheet':
            idx = int(data.get('index', state.active_sheet))
            if len(state.sheets) > 1 and 0 <= idx < len(state.sheets):
                state.sheets.pop(idx)
                state.active_sheet = min(state.active_sheet, len(state.sheets) - 1)
            self._json({'ok': True})

        elif self.path == '/api/github/config':
            gh.token  = data.get('token',  gh.token)
            gh.repo   = data.get('repo',   gh.repo).strip()
            gh.branch = data.get('branch', gh.branch).strip() or 'main'
            gh.path   = data.get('path',   gh.path).strip().lstrip('/')
            save_gh_config()
            self._json({'ok': True})

        elif self.path == '/api/github/commit':
            if not gh.token or not gh.repo or not gh.path:
                self._json({'error': 'GitHub not configured'}); return
            message    = data.get('message', 'Update file').strip() or 'Update file'
            clean_path = gh.path.lstrip('/')
            existing, _ = _gh_api('GET',
                f'/repos/{gh.repo}/contents/{clean_path}?ref={urllib.parse.quote(gh.branch)}')
            file_sha = existing.get('sha') if isinstance(existing, dict) else None
            # Encode content (CSV only — xlsx is handled by /api/github/commit-xlsx)
            if state.filetype == 'xlsx':
                self._json({'error': 'use /api/github/commit-xlsx for xlsx files'}); return
            elif state.filetype == 'json':
                encoded = base64.b64encode(
                    json.dumps(_sheets_to_doc_json(state.sheets), ensure_ascii=False, indent=2).encode('utf-8')
                ).decode()
            else:
                buf = io.StringIO()
                csv.writer(buf).writerow(state.headers)
                csv.writer(buf).writerows(state.rows)
                encoded = base64.b64encode(buf.getvalue().encode('utf-8')).decode()
            payload = {'message': message, 'content': encoded, 'branch': gh.branch}
            if file_sha:
                payload['sha'] = file_sha
            result, err = _gh_api('PUT', f'/repos/{gh.repo}/contents/{clean_path}', payload)
            if err:
                self._json({'error': err})
            else:
                self._json({'ok': True, 'sha': result['commit']['sha'][:7],
                            'url': result['commit']['html_url']})
        else:
            self._send(404, b'Not found')

    # ── helpers ──

    def _serve_html(self):
        raw = _load_html().encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def _json(self, obj):
        raw = json.dumps(obj).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def _send(self, code, body):
        self.send_response(code)
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *_):
        pass   # silence server logs


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    load_gh_config()

    if not _XLSX_OK:
        print('Note: openpyxl not installed — Excel (.xlsx) support disabled.')
        print('      pip install openpyxl\n')

    # Optionally load a file passed as a CLI argument
    if len(sys.argv) > 1:
        path = Path(sys.argv[1]).expanduser().resolve()
        if path.exists():
            if path.suffix.lower() == '.xlsx':
                if not _XLSX_OK:
                    print('Error: openpyxl is required to open .xlsx files.')
                    print('       pip install openpyxl')
                    sys.exit(1)
                state.sheets       = _parse_xlsx_sheets_with_styles(path.read_bytes())
                state.active_sheet = 0
                state.filetype     = 'xlsx'
            elif path.suffix.lower() == '.json':
                state.sheets       = _doc_json_to_sheets(path.read_text(encoding='utf-8'))
                state.active_sheet = 0
                state.filetype     = 'json'
            else:
                with open(path, newline='', encoding='utf-8-sig') as f:
                    rows = list(csv.reader(f))
                if rows:
                    headers = rows[0]
                    parsed_rows = [list(r) for r in rows[1:]]
                    for row in parsed_rows:
                        while len(row) < len(headers):
                            row.append('')
                else:
                    headers = []
                    parsed_rows = []
                state.sheets       = [{'name': 'Sheet1', 'headers': headers, 'rows': parsed_rows}]
                state.active_sheet = 0
                state.filetype     = 'csv'
            state.filepath = str(path)
        else:
            print(f'File not found: {path}')

    server = HTTPServer(('localhost', 0), Handler)
    port   = server.server_address[1]
    url    = f'http://localhost:{port}'

    print(f'CSV Editor  →  {url}')
    print('Press Ctrl+C to quit.\n')

    threading.Timer(0.6, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('Shutting down.')


if __name__ == '__main__':
    main()
